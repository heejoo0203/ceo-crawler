# -*- coding: utf-8 -*-
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# 경로 설정
image_dir = os.path.join(os.getcwd(), "output_photos") #사진이 저장되어 있는 폴더
excel_path = os.path.join(os.getcwd(), "input_ceo_list_원본.xlsx") #사진을 넣을 중복 제거 전 엑셀파일

# 엑셀 불러오기
wb = load_workbook(excel_path)
ws = wb.active

# ------------------------------------------------
# 전체 열 너비 고정 (H열)
# ------------------------------------------------
ws.column_dimensions["H"].width = 12

# personID → row 리스트 매핑 생성
personid_dict = {}

for row in range(2, ws.max_row + 1):
    person_id = ws[f"D{row}"].value  # D열 = personID
    if person_id:
        key = str(person_id).strip()
        if key not in personid_dict:
            personid_dict[key] = []
        personid_dict[key].append(row)

# 이미 존재하는 personID 이미지 목록
image_ids = {
    os.path.splitext(f)[0]
    for f in os.listdir(image_dir)
    if f.lower().endswith(".png")
}

# ------------------------------------------------
# 모든 personID 행에 대해 이미지 또는 "사진 없음" 처리
# ------------------------------------------------
for person_id, rows in personid_dict.items():

    has_image = person_id in image_ids
    img_path = os.path.join(image_dir, person_id + ".png")

    for row in rows:

        # 모든 행 행 높이 고정
        ws.row_dimensions[row].height = 60

        cell = f"H{row}"

        # ------------------------------------------------
        # 사진 있을 경우
        # ------------------------------------------------
        if has_image:
            try:
                img = XLImage(img_path)
                img.width = 80
                img.height = 80
                ws.add_image(img, cell)
                print(f"{person_id}.png → {cell} 삽입 완료")
            except Exception as e:
                print(f"{person_id}.png 삽입 중 오류: {e}")
                ws[cell] = "사진 없음"

        # ------------------------------------------------
        # 사진 없는 경우
        # ------------------------------------------------
        else:
            ws[cell] = "사진 없음"
            print(f"{person_id} → 사진 없음 표시")

# 저장
base, ext = os.path.splitext(excel_path)
output_path = base + "_이미지삽입" + ext
wb.save(output_path)

print(f"모든 작업 완료 ✅ 결과물: {output_path}")
